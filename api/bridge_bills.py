"""ApiBridge domain mixin."""
import json
import logging
import os
import uuid
from datetime import datetime, date, timedelta

from core.db_rebuild import rebuild_database
from core.trade_types import VALID_TRADE_TYPES, TRADE_TYPE_LABELS, get_trade_type_label
from modules.accounting.credit_tracker import CreditTracker
from modules.accounting.cross_platform_merger import CrossPlatformMerger
from modules.accounting.transfer_pairer import TransferPairer

from .bridge_base import DateTimeEncoder, audit_log, logger


class BillsBridgeMixin:
    def query_bills(self, params=None) -> dict:
        p = params or {}
        filters = p.get('filters') or {}
        page = p.get('page', 1)
        page_size = p.get('page_size', 20)

        where, sql_params = self._build_bill_conditions(filters)
        has_merge_filter = filters.get('merge_status')
        join_clause = " LEFT JOIN bill_accounting ba ON ub.id = ba.bill_id " if has_merge_filter else ""
        offset = (page - 1) * page_size

        total = self.dal.fetch_one(
            f"SELECT COUNT(*) as cnt FROM unified_bills ub{join_clause} WHERE {where}",
            tuple(sql_params),
        )['cnt']

        summary_row = self.dal.fetch_one(
            f"SELECT "
            f"COALESCE(SUM(CASE WHEN ub.direction = 'income' THEN ub.amount_cents ELSE 0 END), 0) AS income, "
            f"COALESCE(SUM(CASE WHEN ub.direction = 'expense' THEN ub.amount_cents ELSE 0 END), 0) AS expense "
            f"FROM unified_bills ub{join_clause} WHERE {where}",
            tuple(sql_params),
        )
        summary = {
            'income': summary_row['income'] if summary_row else 0,
            'expense': summary_row['expense'] if summary_row else 0,
        }

        rows = self.dal.fetch_all(
            f"SELECT ub.*, ba.merge_status, ba.transfer_link_id, ba.is_credit, "
            f"ba.credit_account_id, ca.account_name AS credit_account_name, "
            f"bc.name as category_name "
            f"FROM unified_bills ub "
            f"LEFT JOIN bill_accounting ba ON ub.id = ba.bill_id "
            f"LEFT JOIN credit_accounts ca ON ba.credit_account_id = ca.id "
            f"LEFT JOIN bill_categories bc ON ub.category_id = bc.id "
            f"WHERE {where} ORDER BY ub.trade_time DESC LIMIT ? OFFSET ?",
            tuple(sql_params) + (page_size, offset),
        )

        return self.ok({'total': total, 'list': [dict(r) for r in rows], 'summary': summary})

    @audit_log('create_bill')
    def create_bill(self, params=None) -> dict:
        fields = (params or {}).get('fields') or {}
        
        trade_time, err = self._validate_trade_time(fields.get('trade_time'))
        if err:
            return self.err(err)

        trade_type = self._normalize_trade_type(fields.get('trade_type'))
        if trade_type not in VALID_TRADE_TYPES:
            return self.err('交易类型无效')

        direction = fields.get('direction')
        if direction not in ('income', 'expense', 'neutral'):
            return self.err('收支方向无效')

        amount_cents, err = self._validate_amount(fields.get('amount_cents'))
        if err:
            return self.err(err)

        counterparty, err = self._validate_string_field(fields.get('counterparty'), 200, '交易对方')
        if err:
            return self.err(err)

        product_desc, err = self._validate_string_field(fields.get('product_desc'), 500, '商品说明')
        if err:
            return self.err(err)

        remark, err = self._validate_string_field(fields.get('remark'), 1000, '备注')
        if err:
            return self.err(err)

        now = self._now()
        batch_id = f"manual:{uuid.uuid4()}"
        channel_trade_no = fields.get('channel_trade_no') or f"MANUAL_{uuid.uuid4().hex}"

        try:
            with self.dal.transaction():
                existing = self.dal.fetch_one(
                    "SELECT id FROM unified_bills WHERE channel = ? AND channel_trade_no = ?",
                    ('manual', channel_trade_no),
                )
                if existing:
                    return self.err('交易流水号已存在')

                account_id, role_id, assign_status = self._resolve_bill_assignment(
                    fields.get('account_id'), fields.get('role_id')
                )
                category_id = self._validate_existing_id('bill_categories', fields.get('category_id'))
                bill_data = {
                    'channel': 'manual',
                    'trade_time': trade_time,
                    'trade_type': trade_type,
                    'direction': direction,
                    'amount_cents': amount_cents,
                    'counterparty': counterparty or '',
                    'product_desc': product_desc or '',
                    'payment_method': fields.get('payment_method', ''),
                    'status': fields.get('status', ''),
                    'channel_trade_no': channel_trade_no,
                    'remark': remark or '',
                    'account_id': account_id,
                    'role_id': role_id,
                    'category_id': category_id,
                    'assign_status': assign_status,
                    'is_system': 0,
                    'batch_id': batch_id,
                    'is_manual_edited': 1,
                    'created_at': now,
                    'updated_at': now,
                }
                if category_id:
                    bill_data.update({
                        'is_category_manual_edited': 1,
                        'category_source': 'manual',
                        'category_score': 0,
                        'category_rule_id': None,
                    })

                bill_id = self.dal.insert('unified_bills', bill_data)
                imported_bill = {**bill_data, 'id': bill_id}
                self.dal.insert('source_bills', {
                    'bill_id': bill_id,
                    'channel': 'manual',
                    'raw_json': json.dumps({
                        'source': 'manual_create',
                        'version': 1,
                        'fields': fields,
                    }, ensure_ascii=False, cls=DateTimeEncoder),
                    'created_at': now,
                })
                self._apply_accounting_pipeline(imported_bill, now)

                if not category_id:
                    from modules.categorizer import CategoryService
                    category_service = CategoryService(self.dal)
                    category_result = category_service.categorize_bill(bill_id, bill=imported_bill)
                    category_service.apply_result(bill_id, category_result)

            return self.ok({'bill_id': bill_id})
        except Exception as e:
            return self.err(f'新增账单失败: {e}')

    def export_bills(self, params=None) -> dict:
        filters = (params or {}).get('filters') or {}
        where, sql_params = self._build_bill_conditions(filters)
        try:
            import webview
            import openpyxl
            from openpyxl.utils import get_column_letter

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=f"账单导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                file_types=('Excel 文件 (*.xlsx)', '所有文件 (*.*)'),
            )
            if not result:
                return self.ok({'path': '', 'count': 0, 'cancelled': True}, '已取消导出')
            export_path = result if isinstance(result, str) else result[0]
            if not export_path.lower().endswith('.xlsx'):
                export_path += '.xlsx'

            rows = self.dal.fetch_all(
                f"SELECT ub.*, ba.merge_status, ba.transfer_link_id, ba.is_credit, "
                f"bc.name AS category_name, r.name AS role_name, a.account_name "
                f"FROM unified_bills ub "
                f"LEFT JOIN bill_accounting ba ON ub.id = ba.bill_id "
                f"LEFT JOIN bill_categories bc ON ub.category_id = bc.id "
                f"LEFT JOIN roles r ON ub.role_id = r.id "
                f"LEFT JOIN accounts a ON ub.account_id = a.id "
                f"WHERE {where} ORDER BY ub.trade_time DESC",
                tuple(sql_params),
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'bills'
            headers = [
                'channel', 'trade_time', 'trade_type', 'direction', 'amount_cents', 'amount_yuan',
                'counterparty', 'product_desc', 'payment_method', 'status', 'channel_trade_no', 'remark',
                'account_id', 'account_name', 'role_id', 'role_name', 'category_id', 'category_name',
                'assign_status', 'is_manual_edited', 'is_category_manual_edited', 'category_source',
                'is_system', 'merge_status', 'channel_label', 'trade_type_label', 'direction_label',
            ]
            ws.append(headers)
            channel_labels = {'wechat': '微信', 'alipay': '支付宝', 'ccb': '建行', 'manual': '手工记账'}
            direction_labels = {'income': '收入', 'expense': '支出', 'neutral': '中性'}
            for row in rows:
                d = dict(row)
                ws.append([
                    d.get('channel'), d.get('trade_time'), d.get('trade_type'), d.get('direction'),
                    d.get('amount_cents'), (d.get('amount_cents') or 0) / 100,
                    d.get('counterparty'), d.get('product_desc'), d.get('payment_method'), d.get('status'),
                    d.get('channel_trade_no'), d.get('remark'), d.get('account_id'), d.get('account_name'),
                    d.get('role_id'), d.get('role_name'), d.get('category_id'), d.get('category_name'),
                    d.get('assign_status'), d.get('is_manual_edited'), d.get('is_category_manual_edited'),
                    d.get('category_source'), d.get('is_system'), d.get('merge_status'),
                    channel_labels.get(d.get('channel'), d.get('channel')),
                    get_trade_type_label(d.get('trade_type')),
                    direction_labels.get(d.get('direction'), d.get('direction')),
                ])
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16

            meta = wb.create_sheet('__meta__')
            meta.append(['key', 'value'])
            meta.append(['format', 'all_wallet_manager_export'])
            meta.append(['version', '1'])
            meta.append(['exported_at', self._now()])
            meta.append(['filters_json', json.dumps(filters, ensure_ascii=False, cls=DateTimeEncoder)])
            wb.save(export_path)
            wb.close()
            return self.ok({'path': export_path, 'count': len(rows)})
        except Exception as e:
            return self.err(f'导出失败: {e}')

    def get_bill_detail(self, params=None) -> dict:
        bill_id = (params or {}).get('bill_id')
        bill = self.dal.fetch_one(
            "SELECT ub.*, ba.merge_status, ba.transfer_link_id, "
            "ba.is_credit, ba.credit_account_id, ca.account_name AS credit_account_name, "
            "ba.merged_group_id, ba.real_payer_account_id "
            "FROM unified_bills ub "
            "LEFT JOIN bill_accounting ba ON ub.id = ba.bill_id "
            "LEFT JOIN credit_accounts ca ON ba.credit_account_id = ca.id "
            "WHERE ub.id = ?",
            (bill_id,),
        )
        if not bill:
            return self.err('账单不存在')

        source = self.dal.fetch_one(
            "SELECT * FROM source_bills WHERE bill_id = ?", (bill_id,)
        )

        result = dict(bill)
        if source:
            result['source_bill'] = {
                'channel': source['channel'],
                'raw_json': source['raw_json'],
            }

        return self.ok(result)

    @audit_log('update_bill')
    def update_bill(self, params=None) -> dict:
        p = params or {}
        bill_id = p.get('bill_id')
        fields = p.get('fields', {})
        bill = self.dal.fetch_one(
            "SELECT is_system FROM unified_bills WHERE id = ?", (bill_id,)
        )
        if not bill:
            return self.err('账单不存在')
        if bill['is_system'] == 1:
            return self.err('系统生成的记录不可编辑')

        allowed = {
            'counterparty', 'product_desc', 'payment_method', 'remark',
            'trade_type', 'direction', 'category_id', 'assign_status',
            'account_id', 'role_id',
        }
        data = {k: v for k, v in fields.items() if k in allowed}
        if not data:
            return self.err('无有效更新字段')

        if 'trade_type' in data:
            trade_type = self._normalize_trade_type(data.get('trade_type'))
            if trade_type is None:
                data['trade_type'] = None
            elif trade_type not in VALID_TRADE_TYPES:
                return self.err('交易类型无效')
            else:
                data['trade_type'] = trade_type

        if 'direction' in data and data.get('direction') not in ('income', 'expense', 'neutral'):
            return self.err('收支方向无效')

        if 'category_id' in data:
            data['is_category_manual_edited'] = 1
            data['category_source'] = 'manual'
            data['category_score'] = 0
            data['category_rule_id'] = None
        data['is_manual_edited'] = 1
        data['updated_at'] = self._now()

        self.dal.update('unified_bills', data, 'id = ?', (bill_id,))
        return self.ok()

    def batch_update_bills(self, params=None) -> dict:
        p = params or {}
        bill_ids = p.get('bill_ids', [])
        fields = p.get('fields', {})
        if not bill_ids:
            return self.err('未指定账单')

        allowed = {
            'category_id', 'assign_status', 'account_id', 'role_id',
        }
        data = {k: v for k, v in fields.items() if k in allowed}
        if not data:
            return self.err('无有效更新字段')

        if 'category_id' in data:
            data['is_category_manual_edited'] = 1
            data['category_source'] = 'manual'
            data['category_score'] = 0
            data['category_rule_id'] = None
        data['updated_at'] = self._now()
        placeholders = ', '.join(['?' for _ in bill_ids])
        updated = self.dal.update(
            'unified_bills',
            data,
            f'id IN ({placeholders})',
            tuple(bill_ids),
        )
        return self.ok({'updated': updated})
