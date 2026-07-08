"""
自导出账单解析器 —— 解析本工具导出的 Excel 格式
"""
import os

from .base_parser import BaseParser, ParseResult


class SelfExportParser(BaseParser):
    FORMAT_NAME = 'all_wallet_manager_export'
    VERSION = '1'
    REQUIRED_HEADERS = {
        'channel', 'trade_time', 'trade_type', 'direction',
        'amount_cents', 'channel_trade_no',
    }

    def __init__(self, config_manager=None):
        self.config_manager = config_manager

    def get_channel(self) -> str:
        return 'self_export'

    def parse(self, file_path: str) -> ParseResult:
        ext = os.path.splitext(file_path)[1].lower()
        if ext != '.xlsx':
            return ParseResult(errors=[f'自导出格式仅支持 .xlsx 文件: {ext}'])

        import openpyxl

        result = ParseResult()
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            if '__meta__' not in wb.sheetnames or 'bills' not in wb.sheetnames:
                result.errors.append('不是有效的自导出账单文件')
                return result

            meta = self._read_meta(wb['__meta__'])
            if meta.get('format') != self.FORMAT_NAME:
                result.errors.append('自导出格式标识不匹配')
                return result
            if str(meta.get('version')) != self.VERSION:
                result.errors.append(f"不支持的自导出格式版本: {meta.get('version')}")
                return result

            ws = wb['bills']
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return result

            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            missing = self.REQUIRED_HEADERS - set(headers)
            if missing:
                result.errors.append(f"自导出文件缺少列: {', '.join(sorted(missing))}")
                return result

            for row in rows[1:]:
                if not row or all(v is None for v in row):
                    continue
                row_dict = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        row_dict[header] = row[i]
                record = self._standardize(row_dict)
                if record:
                    result.add_record(record, {'channel': 'self_export', 'raw': row_dict})

            result.total = result.success + result.duplicate
            return result
        except Exception as e:
            result.errors.append(f'解析自导出文件失败: {e}')
            return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _read_meta(self, ws) -> dict:
        meta = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            meta[str(row[0])] = '' if len(row) < 2 or row[1] is None else str(row[1])
        return meta

    def _standardize(self, row: dict) -> dict | None:
        channel = str(row.get('channel') or '').strip()
        trade_no = str(row.get('channel_trade_no') or '').strip()
        if not channel or not trade_no:
            return None

        return {
            'channel': channel,
            'trade_time': str(row.get('trade_time') or '').strip(),
            'trade_type': str(row.get('trade_type') or '').strip(),
            'direction': str(row.get('direction') or '').strip(),
            'amount_cents': int(float(row.get('amount_cents') or 0)),
            'counterparty': row.get('counterparty') or '',
            'product_desc': row.get('product_desc') or '',
            'payment_method': row.get('payment_method') or '',
            'status': row.get('status') or '',
            'channel_trade_no': trade_no,
            'remark': row.get('remark') or '',
            'account_id': self._int_or_none(row.get('account_id')),
            'role_id': self._int_or_none(row.get('role_id')),
            'category_id': self._int_or_none(row.get('category_id')),
            'assign_status': row.get('assign_status') or '',
            'is_manual_edited': self._int_or_zero(row.get('is_manual_edited')),
            'is_category_manual_edited': self._int_or_zero(row.get('is_category_manual_edited')),
            'category_source': row.get('category_source') or 'auto',
            'is_system': self._int_or_zero(row.get('is_system')),
        }

    def _int_or_none(self, value):
        if value in (None, ''):
            return None
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    def _int_or_zero(self, value) -> int:
        parsed = self._int_or_none(value)
        return parsed if parsed is not None else 0
