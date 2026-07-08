"""
账户标识提取器 —— 从账单文件/源数据中提取账户标识信息
"""
import os
import re
import openpyxl


class AccountExtractor:
    """从各渠道账单中提取账户标识"""

    # 支付方式分类关键词
    WECHAT_BANK_KEYWORDS = ['银行', '储蓄卡', '信用卡']  # 银行卡类
    WECHAT_CREDIT_KEYWORDS = ['分付']  # 贷款类
    WECHAT_BALANCE_KEYWORDS = ['零钱', '零钱通']  # 余额类
    WECHAT_FAMILY_KEYWORDS = ['亲属卡']  # 亲属卡
    WECHAT_DISCOUNT_KEYWORDS = ['立减', '红包', '优惠', '券', '随机']  # 非真实付款

    ALIPAY_BANK_KEYWORDS = ['银行', '储蓄卡', '信用卡']  # 银行卡类
    ALIPAY_CREDIT_KEYWORDS = ['花呗', '信用支付', '贷']  # 贷款类
    ALIPAY_BALANCE_KEYWORDS = ['余额', '余额宝']  # 余额类
    ALIPAY_DISCOUNT_KEYWORDS = ['立减', '红包', '优惠', '券', '随机']  # 非真实付款

    # ─── 微信账户提取 ─────────────────────────────────

    def extract_wechat_account_info(self, file_path: str) -> dict:
        """
        从微信账单元信息中提取账户标识
        返回: { 'nickname': '微信昵称', 'accounts': [{'tag': '账户标识', 'name': '账户名', 'payment_method': '支付方式'}] }
        """
        nickname = None
        accounts = []

        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xlsx':
            nickname = self._extract_wechat_nickname_xlsx(file_path)
        elif ext == '.csv':
            nickname = self._extract_wechat_nickname_csv(file_path)

        if not nickname:
            nickname = '未知用户'

        # 从账单数据中提取所有支付方式
        payment_methods = self._extract_wechat_payment_methods(file_path)

        # 分类处理支付方式
        has_bank_account = False
        bank_accounts = []
        credit_accounts = []
        balance_accounts = []
        family_accounts = []
        other_accounts = []

        for pm in payment_methods:
            pm_type = self._classify_wechat_payment_method(pm)

            alias_candidates = [nickname] if nickname and nickname != '未知用户' else []

            if pm_type == 'bank':
                card_suffix = self._extract_card_suffix(pm)
                bank_name = self._extract_bank_name(pm)
                tag = f"wechat-{nickname}-{bank_name}_{card_suffix}" if bank_name and card_suffix else f"wechat-{nickname}-{card_suffix}"
                name = f"微信-{nickname}-{pm}"
                bank_accounts.append({
                    'tag': tag,
                    'name': name,
                    'payment_method': pm,
                    'payment_method_type': pm_type,
                    'card_suffix': card_suffix,
                    'bank_name': bank_name,
                    'alias_candidates': alias_candidates,
                })
                has_bank_account = True
            elif pm_type == 'credit':
                tag = f"wechat-{nickname}-{pm}"
                name = f"微信-{nickname}-{pm}"
                credit_accounts.append({
                    'tag': tag,
                    'name': name,
                    'payment_method': pm,
                    'payment_method_type': pm_type,
                    'alias_candidates': alias_candidates,
                })
            elif pm_type == 'balance':
                tag = f"wechat-{nickname}-{pm}"
                name = f"微信-{nickname}-{pm}"
                balance_accounts.append({
                    'tag': tag,
                    'name': name,
                    'payment_method': pm,
                    'payment_method_type': pm_type,
                    'alias_candidates': alias_candidates,
                })
            elif pm_type == 'family':
                family_payer = self._extract_family_card_name(pm)
                tag = f"wechat-{nickname}-{pm}"
                name = f"微信-{nickname}-{pm}"
                family_accounts.append({
                    'tag': tag,
                    'name': name,
                    'payment_method': pm,
                    'payment_method_type': pm_type,
                    'family_card_payer_name': family_payer,
                    'alias_candidates': [v for v in (nickname, family_payer) if v and v != '未知用户'],
                })
            elif pm_type == 'discount':
                # 非真实付款统一一个账户
                pass  # 最后统一添加
            else:
                # 未识别的归入其他
                pass  # 最后统一添加

        # 合并账户列表
        accounts = bank_accounts + credit_accounts + balance_accounts + family_accounts

        # 非真实付款统一账户
        accounts.append({
            'tag': f"wechat-{nickname}-其他",
            'name': f"微信-{nickname}-其他优惠",
            'payment_method': '_discount_',
            'payment_method_type': 'discount',
            'alias_candidates': [nickname] if nickname and nickname != '未知用户' else [],
        })

        return {
            'nickname': nickname,
            'accounts': accounts,
        }

    def _classify_wechat_payment_method(self, pm: str) -> str:
        """分类微信支付方式"""
        pm_lower = pm.lower()
        for kw in self.WECHAT_DISCOUNT_KEYWORDS:
            if kw in pm:
                return 'discount'
        for kw in self.WECHAT_BANK_KEYWORDS:
            if kw in pm:
                return 'bank'
        for kw in self.WECHAT_CREDIT_KEYWORDS:
            if kw in pm:
                return 'credit'
        for kw in self.WECHAT_FAMILY_KEYWORDS:
            if kw in pm:
                return 'family'
        for kw in self.WECHAT_BALANCE_KEYWORDS:
            if kw in pm:
                return 'balance'
        return 'other'

    def _extract_wechat_nickname_xlsx(self, file_path: str) -> str:
        """从xlsx元信息提取微信昵称"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            # 第1行: "微信昵称：[风]"
            row1 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            wb.close()

            if row1 and row1[0]:
                text = str(row1[0])
                match = re.search(r'微信昵称：\[([^\]]+)\]', text)
                if match:
                    return match.group(1)
        except Exception:
            pass
        return '未知用户'

    def _extract_wechat_nickname_csv(self, file_path: str) -> str:
        """从csv元信息提取微信昵称"""
        for encoding in ('utf-8-sig', 'gbk', 'utf-8'):
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    lines = f.readlines()
                    for line in lines[:5]:
                        match = re.search(r'微信昵称：\[([^\]]+)\]', line)
                        if match:
                            return match.group(1)
            except (UnicodeDecodeError, UnicodeError):
                continue
        return '未知用户'

    def _extract_wechat_payment_methods(self, file_path: str) -> list[str]:
        """从微信账单数据中提取所有支付方式"""
        payment_methods = set()
        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i < 17:  # 跳过元信息和表头
                        continue
                    if row and len(row) >= 7:
                        pm = row[6]  # 支付方式列
                        if pm:
                            payment_methods.add(str(pm).strip())
                wb.close()
            elif ext == '.csv':
                for encoding in ('utf-8-sig', 'gbk', 'utf-8'):
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            for line in lines[17:]:  # 跳过元信息和表头
                                parts = self._split_csv_line(line.strip())
                                if len(parts) >= 7:
                                    pm = parts[6]
                                    if pm:
                                        payment_methods.add(pm.strip())
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
        except Exception:
            pass

        return list(payment_methods)

    # ─── 支付宝账户提取 ─────────────────────────────────

    def extract_alipay_account_info(self, file_path: str) -> dict:
        """
        从支付宝账单中提取账户标识
        返回: { 'account_suffix': '账号尾号', 'accounts': [{'tag': '账户标识', 'name': '账户名', 'payment_method': '支付方式'}] }
        """
        accounts = []

        # 提取支付宝账号尾号
        account_suffix = self._extract_alipay_account_suffix(file_path)
        if not account_suffix:
            account_suffix = '未知'

        payment_methods = self._extract_alipay_payment_methods(file_path)

        # 分类处理支付方式（使用集合去重）
        bank_account_tags = set()
        credit_account_tags = set()
        balance_account_tags = set()
        bank_accounts = []
        credit_accounts = []
        balance_accounts = []
        has_other_discount = False

        for pm in payment_methods:
            pm_type = self._classify_alipay_payment_method(pm)
            pm_normalized = self._normalize_alipay_payment_method(pm)

            if pm_type == 'bank':
                card_suffix = self._extract_card_suffix(pm)
                bank_name = self._extract_bank_name(pm)
                tag = f"alipay-{account_suffix}-{bank_name}_{card_suffix}" if bank_name and card_suffix else f"alipay-{account_suffix}-{card_suffix}"
                if tag not in bank_account_tags:
                    bank_account_tags.add(tag)
                    name = f"支付宝-{account_suffix}-{pm_normalized}"
                    bank_accounts.append({
                        'tag': tag,
                        'name': name,
                        'payment_method': pm_normalized,
                        'payment_method_type': pm_type,
                    })
            elif pm_type == 'credit':
                tag = f"alipay-{account_suffix}-{pm_normalized}"
                if tag not in credit_account_tags:
                    credit_account_tags.add(tag)
                    name = f"支付宝-{account_suffix}-{pm_normalized}"
                    credit_accounts.append({
                        'tag': tag,
                        'name': name,
                        'payment_method': pm_normalized,
                        'payment_method_type': pm_type,
                    })
            elif pm_type == 'balance':
                tag = f"alipay-{account_suffix}-{pm_normalized}"
                if tag not in balance_account_tags:
                    balance_account_tags.add(tag)
                    name = f"支付宝-{account_suffix}-{pm_normalized}"
                    balance_accounts.append({
                        'tag': tag,
                        'name': name,
                        'payment_method': pm_normalized,
                        'payment_method_type': pm_type,
                    })
            elif pm_type == 'discount':
                # 非真实付款统一一个账户
                has_other_discount = True
            else:
                # 未识别的归入其他
                pass

        # 合并账户列表
        accounts = bank_accounts + credit_accounts + balance_accounts

        # 非真实付款统一账户
        if has_other_discount:
            accounts.append({
                'tag': f"alipay-{account_suffix}-优惠",
                'name': f"支付宝-{account_suffix}-优惠",
                'payment_method': '_discount_',
                'payment_method_type': 'discount',
            })

        return {'account_suffix': account_suffix, 'accounts': accounts}

    def _classify_alipay_payment_method(self, pm: str) -> str:
        """分类支付宝支付方式"""
        for kw in self.ALIPAY_DISCOUNT_KEYWORDS:
            if kw in pm:
                return 'discount'
        for kw in self.ALIPAY_BANK_KEYWORDS:
            if kw in pm:
                return 'bank'
        for kw in self.ALIPAY_CREDIT_KEYWORDS:
            if kw in pm:
                return 'credit'
        for kw in self.ALIPAY_BALANCE_KEYWORDS:
            if kw in pm:
                return 'balance'
        return 'other'

    def _normalize_alipay_payment_method(self, pm: str) -> str:
        """规范化支付宝支付方式名称，提取核心部分"""
        # 账户余额类：提取"账户余额"或"余额宝"
        if '账户余额' in pm:
            return '账户余额'
        if '余额宝' in pm:
            return '余额宝'
        # 花呗类
        if '花呗' in pm:
            return '花呗'
        # 银行卡类：保留原格式
        if any(kw in pm for kw in self.ALIPAY_BANK_KEYWORDS):
            return pm
        return pm

    def _extract_alipay_account_suffix(self, file_path: str) -> str:
        """从支付宝账单元信息提取账号尾号"""
        ext = os.path.splitext(file_path)[1].lower()
        suffix = None

        try:
            if ext == '.csv':
                for encoding in ('gbk', 'utf-8-sig', 'utf-8', 'gb18030'):
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            # 查找元信息中的账号（通常在前几行）
                            for line in lines[:25]:
                                # 匹配 "账号：xxx" 或 "支付宝账号：xxx"
                                match = re.search(r'账号[：:]\s*(\S+)', line)
                                if match:
                                    account_str = match.group(1)
                                    # 提取手机号或邮箱尾号
                                    digits = re.sub(r'\D', '', account_str)
                                    if len(digits) >= 4:
                                        suffix = digits[-4:]
                                        break
                                    # 如果是邮箱格式，取@前的最后4位
                                    if '@' in account_str:
                                        email_prefix = account_str.split('@')[0]
                                        suffix = email_prefix[-4:] if len(email_prefix) >= 4 else email_prefix
                                        break
                        if suffix:
                            break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
            elif ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 25:  # 只检查前25行元信息
                        break
                    for cell in row:
                        if cell and isinstance(cell, str):
                            match = re.search(r'账号[：:]\s*(\S+)', cell)
                            if match:
                                account_str = match.group(1)
                                digits = re.sub(r'\D', '', account_str)
                                if len(digits) >= 4:
                                    suffix = digits[-4:]
                                    break
                                if '@' in account_str:
                                    email_prefix = account_str.split('@')[0]
                                    suffix = email_prefix[-4:] if len(email_prefix) >= 4 else email_prefix
                                    break
                wb.close()
        except Exception:
            pass

        return suffix

    def _extract_alipay_payment_methods(self, file_path: str) -> list[str]:
        """从支付宝账单数据中提取所有支付方式"""
        payment_methods = set()
        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == '.csv':
                for encoding in ('gbk', 'utf-8-sig', 'utf-8', 'gb18030'):
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            lines = f.readlines()
                            header_idx = -1
                            for i, line in enumerate(lines):
                                if '交易时间' in line and '交易分类' in line:
                                    header_idx = i
                                    break
                            if header_idx >= 0:
                                for line in lines[header_idx + 1:]:
                                    parts = self._split_csv_line(line.strip())
                                    if len(parts) >= 8:
                                        pm = parts[7]  # 收/付款方式列
                                        if pm:
                                            payment_methods.add(pm.strip())
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
            elif ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                header_idx = -1
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if row and '交易时间' in str(row[0] or ''):
                        header_idx = i
                        break
                if header_idx >= 0:
                    for row in list(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
                        if row and len(row) >= 8:
                            pm = row[7]
                            if pm:
                                payment_methods.add(str(pm).strip())
                wb.close()
        except Exception:
            pass

        return list(payment_methods)

    # ─── 建行账户提取 ─────────────────────────────────

    def extract_ccb_account_info(self, file_path: str) -> dict:
        """
        从建行账单中提取账户标识
        - xls: 从元信息行提取账号后4位
        - pdf: 从页面文本提取账号后4位
        """
        ext = os.path.splitext(file_path)[1].lower()
        account_tag = None

        if ext == '.xls':
            account_tag = self._extract_ccb_tag_xls(file_path)
        elif ext == '.pdf':
            account_tag = self._extract_ccb_tag_pdf(file_path)

        if account_tag:
            return {
                'accounts': [{
                    'tag': f'ccb-{account_tag}',
                    'name': f'建行卡({account_tag})',
                    'payment_method': '',
                }],
                'need_manual_assign': False,
            }

        return {
            'accounts': [],
            'need_manual_assign': True,
        }

    def _extract_ccb_tag_xls(self, file_path: str) -> str:
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            # 查找包含"账号"的单元格（不查找"账户"，避免混淆）
            for r in range(min(ws.nrows, 10)):
                for c in range(ws.ncols):
                    val = str(ws.cell_value(r, c)).strip()
                    if '账号' in val and '客户' not in val:  # 排除"客户账号"等非银行卡号
                        # 优先从当前单元格提取（标签和值可能在同一合并单元格中）
                        digits = re.sub(r'\D', '', val)
                        if len(digits) >= 4:
                            return digits[-4:]
                        # 再从右侧单元格查找
                        for c2 in range(c + 1, ws.ncols):
                            v = str(ws.cell_value(r, c2)).strip()
                            if v and v != '账号':
                                # 跳过日期/币种/金额类单元格，避免误提取
                                if any(kw in v for kw in ('日期', '时间', '币种', '金额', '余额')):
                                    continue
                                digits = re.sub(r'\D', '', v)
                                if len(digits) >= 4:
                                    return digits[-4:]
                        # 也检查下一行
                        if r + 1 < ws.nrows:
                            for c2 in range(ws.ncols):
                                v = str(ws.cell_value(r + 1, c2)).strip()
                                if v:
                                    # 跳过日期/币种/金额类单元格，避免误提取
                                    if any(kw in v for kw in ('日期', '时间', '币种', '金额', '余额')):
                                        continue
                                    digits = re.sub(r'\D', '', v)
                                    if len(digits) >= 4:
                                        return digits[-4:]
        except Exception:
            pass
        return None

    def _extract_ccb_tag_pdf(self, file_path: str) -> str:
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text() or ''
                    # 匹配银行卡号（16-19位数字）
                    match = re.search(r'账号[：:]\s*(\d{16,19})', text)
                    if match:
                        return match.group(1)[-4:]
                    # 匹配"卡号"关键词
                    match = re.search(r'卡号[：:]\s*(\d{16,19})', text)
                    if match:
                        return match.group(1)[-4:]
        except Exception:
            pass
        return None

    # ─── 通用辅助 ─────────────────────────────────

    def _extract_card_suffix(self, payment_method: str) -> str:
        """从支付方式中提取银行卡尾号（如(0480)）"""
        match = re.search(r'\(([0-9]{4})\)', payment_method)
        if match:
            return match.group(1)
        return ''

    def _extract_bank_name(self, payment_method: str) -> str:
        """从支付方式中提取银行名称（如"招商银行储蓄卡(1234)" -> "招商银行"）"""
        # 移除尾号部分
        pm_clean = re.sub(r'\([0-9]{4}\)', '', payment_method)
        # 匹配银行名
        match = re.match(r'^([^\s]+银行)', pm_clean)
        if match:
            return match.group(1)
        return ''

    def _extract_family_card_name(self, payment_method: str) -> str:
        """从亲属卡支付方式中尽量提取代付者/持卡人昵称。"""
        text = str(payment_method or '').strip()
        if not text or '亲属卡' not in text:
            return ''

        patterns = (
            r'([^\s（）()，,：:]+)的亲属卡',
            r'亲属卡[（(]([^）)]+)[）)]',
            r'亲属卡[：:]\s*([^\s，,]+)',
            r'亲属卡-([^\s，,]+)',
        )
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                name = match.group(1).strip()
                if name and name not in ('亲属卡', '支付'):
                    return name
        return ''

    def _clean_payment_method(self, payment_method: str) -> str:
        """清理支付方式名称，移除括号内容"""
        return re.sub(r'\([^)]*\)', '', payment_method).strip()

    def _split_csv_line(self, line: str) -> list[str]:
        """解析CSV行"""
        result = []
        current = []
        in_quotes = False
        for ch in line:
            if ch == '"':
                in_quotes = not in_quotes
            elif ch == ',' and not in_quotes:
                result.append(''.join(current).strip())
                current = []
            else:
                current.append(ch)
        result.append(''.join(current).strip())
        return result

    # ─── 统一入口 ─────────────────────────────────

    def extract(self, channel: str, file_path: str) -> dict:
        """根据渠道提取账户信息"""
        if channel == 'wechat':
            return self.extract_wechat_account_info(file_path)
        elif channel == 'alipay':
            return self.extract_alipay_account_info(file_path)
        elif channel == 'ccb':
            return self.extract_ccb_account_info(file_path)
        return {'accounts': []}